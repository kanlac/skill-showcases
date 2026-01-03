#!/usr/bin/env python3
"""
销售数据分析脚本
读取销售数据并填充月度汇总和产品分析表
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict

# 读取文件
wb = openpyxl.load_workbook('output/result_without_skill_2.xlsx')

# 获取各个工作表
sales_sheet = wb['原始销售数据']
product_sheet = wb['产品主数据']
monthly_sheet = wb['月度汇总']
product_analysis_sheet = wb['产品分析']

print("开始数据分析...")

# ===== 1. 处理月度汇总表 =====
print("\n1. 分析月度数据...")

# 从原始数据中提取月度汇总
monthly_data = defaultdict(lambda: {'sales': 0, 'qty': 0, 'total_amount': 0})

# 遍历原始销售数据（从第3行开始，第1行是标题，第2行是列名）
for row in range(3, sales_sheet.max_row + 1):
    year_month = sales_sheet.cell(row, 1).value  # 年月
    qty = sales_sheet.cell(row, 5).value  # 销售数量
    amount = sales_sheet.cell(row, 7).value  # 销售额

    if year_month and qty and amount:
        monthly_data[year_month]['sales'] += float(amount)
        monthly_data[year_month]['qty'] += int(qty)
        monthly_data[year_month]['total_amount'] += float(amount)

# 获取2023和2024年的月度目标
# 从区域目标表读取年度目标，然后平均到每个月
region_sheet = wb['区域目标']
target_2023 = 0
target_2024 = 0

for row in range(2, region_sheet.max_row + 1):
    target_2023 += float(region_sheet.cell(row, 2).value or 0)
    target_2024 += float(region_sheet.cell(row, 3).value or 0)

# 月度目标 = 年度目标 / 12
monthly_target_2023 = target_2023 / 12
monthly_target_2024 = target_2024 / 12

print(f"2023年月度目标: {monthly_target_2023:,.2f}")
print(f"2024年月度目标: {monthly_target_2024:,.2f}")

# 填充月度汇总表（从第3行开始）
sorted_months = sorted(monthly_data.keys())
prev_sales = None

# 红色填充（完成率低于90%）
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

for idx, year_month in enumerate(sorted_months):
    row_idx = idx + 3  # 从第3行开始填充
    data = monthly_data[year_month]

    # 总销售额
    total_sales = data['sales']
    monthly_sheet.cell(row_idx, 2).value = round(total_sales, 2)

    # 总销售数量
    total_qty = data['qty']
    monthly_sheet.cell(row_idx, 3).value = total_qty

    # 平均单价
    avg_price = total_sales / total_qty if total_qty > 0 else 0
    monthly_sheet.cell(row_idx, 4).value = round(avg_price, 2)

    # 同比增长率（需要去年同月数据）
    year = int(year_month.split('-')[0])
    month = int(year_month.split('-')[1])
    last_year_month = f"{year-1}-{month:02d}"

    if last_year_month in monthly_data:
        last_year_sales = monthly_data[last_year_month]['sales']
        yoy_growth = (total_sales - last_year_sales) / last_year_sales if last_year_sales > 0 else 0
        monthly_sheet.cell(row_idx, 5).value = round(yoy_growth, 4)  # 保留4位小数，显示为百分比
    else:
        monthly_sheet.cell(row_idx, 5).value = None

    # 环比增长率
    if prev_sales is not None:
        mom_growth = (total_sales - prev_sales) / prev_sales if prev_sales > 0 else 0
        monthly_sheet.cell(row_idx, 6).value = round(mom_growth, 4)
    else:
        monthly_sheet.cell(row_idx, 6).value = None

    prev_sales = total_sales

    # 目标完成率
    target = monthly_target_2023 if year == 2023 else monthly_target_2024
    completion_rate = total_sales / target if target > 0 else 0
    monthly_sheet.cell(row_idx, 7).value = round(completion_rate, 4)

    # 如果完成率低于90%，标红
    if completion_rate < 0.9:
        for col in range(1, 9):  # 标红整行
            monthly_sheet.cell(row_idx, col).fill = red_fill
        monthly_sheet.cell(row_idx, 8).value = "完成率偏低"
    elif completion_rate >= 1.1:
        monthly_sheet.cell(row_idx, 8).value = "超额完成"
    else:
        monthly_sheet.cell(row_idx, 8).value = "正常"

print(f"月度汇总表填充完成，共 {len(sorted_months)} 个月")

# ===== 2. 处理产品分析表 =====
print("\n2. 分析产品数据...")

# 从原始数据中按产品和年份汇总
product_yearly_data = defaultdict(lambda: defaultdict(lambda: {'sales': 0, 'qty': 0, 'cost_total': 0}))

# 先读取产品成本数据
product_costs = {}
for row in range(2, product_sheet.max_row + 1):
    product_name = product_sheet.cell(row, 1).value
    standard_cost = float(product_sheet.cell(row, 4).value or 0)
    target_margin = float(product_sheet.cell(row, 5).value or 0)
    product_costs[product_name] = {
        'cost': standard_cost,
        'target_margin': target_margin
    }

# 从原始销售数据中提取产品数据
for row in range(3, sales_sheet.max_row + 1):
    year_month = sales_sheet.cell(row, 1).value
    product_name = sales_sheet.cell(row, 2).value
    qty = sales_sheet.cell(row, 5).value
    amount = sales_sheet.cell(row, 7).value

    if year_month and product_name and qty and amount:
        year = year_month.split('-')[0]
        qty = int(qty)
        amount = float(amount)

        product_yearly_data[product_name][year]['sales'] += amount
        product_yearly_data[product_name][year]['qty'] += qty

        # 计算成本
        if product_name in product_costs:
            product_yearly_data[product_name][year]['cost_total'] += qty * product_costs[product_name]['cost']

# 计算总销售额用于占比
total_sales_all = sum(
    data['2023']['sales'] + data['2024']['sales']
    for data in product_yearly_data.values()
)

# 填充产品分析表（从第3行开始）
product_list = []

for product_name in sorted(product_yearly_data.keys()):
    data = product_yearly_data[product_name]

    sales_2023 = data['2023']['sales']
    sales_2024 = data['2024']['sales']
    cost_total = data['2023']['cost_total'] + data['2024']['cost_total']
    total_sales = sales_2023 + sales_2024

    # 增长率
    growth_rate = (sales_2024 - sales_2023) / sales_2023 if sales_2023 > 0 else 0

    # 实际毛利率
    margin_amount = total_sales - cost_total
    actual_margin_rate = margin_amount / total_sales if total_sales > 0 else 0

    # 销售占比
    sales_ratio = total_sales / total_sales_all if total_sales_all > 0 else 0

    product_list.append({
        'name': product_name,
        'sales_2023': sales_2023,
        'sales_2024': sales_2024,
        'growth_rate': growth_rate,
        'actual_margin_rate': actual_margin_rate,
        'margin_amount': margin_amount,
        'sales_ratio': sales_ratio,
        'total_sales': total_sales
    })

# 按总销售额排序（用于ABC分类）
product_list.sort(key=lambda x: x['total_sales'], reverse=True)

# ABC分类：累计销售占比0-70%为A类，70-90%为B类，90-100%为C类
cumulative_ratio = 0
for product in product_list:
    cumulative_ratio += product['sales_ratio']

    if cumulative_ratio <= 0.7:
        product['abc'] = 'A类'
    elif cumulative_ratio <= 0.9:
        product['abc'] = 'B类'
    else:
        product['abc'] = 'C类'

# 生成建议
for product in product_list:
    suggestions = []

    # 获取目标毛利率
    target_margin = product_costs.get(product['name'], {}).get('target_margin', 0.3)

    # 根据ABC分类和表现给出建议
    if product['abc'] == 'A类':
        suggestions.append("核心产品")
        if product['growth_rate'] > 0.1:
            suggestions.append("增长良好，继续投入")
        elif product['growth_rate'] < 0:
            suggestions.append("增长放缓，需要关注")
        else:
            suggestions.append("稳定增长")

        if product['actual_margin_rate'] < target_margin:
            suggestions.append("毛利率偏低，优化成本")

    elif product['abc'] == 'B类':
        suggestions.append("重要产品")
        if product['growth_rate'] > 0.15:
            suggestions.append("高增长，有潜力升为A类")
        elif product['growth_rate'] < -0.05:
            suggestions.append("下滑明显，评估是否继续")

        if product['actual_margin_rate'] < target_margin:
            suggestions.append("提升毛利率")

    else:  # C类
        suggestions.append("长尾产品")
        if product['growth_rate'] < -0.1:
            suggestions.append("考虑逐步淘汰")
        elif product['actual_margin_rate'] >= target_margin:
            suggestions.append("毛利率尚可，保持现状")
        else:
            suggestions.append("低优先级")

    product['suggestion'] = "；".join(suggestions)

# 填充到表格
for idx, product in enumerate(product_list):
    row_idx = idx + 3

    product_analysis_sheet.cell(row_idx, 2).value = round(product['sales_2023'], 2)
    product_analysis_sheet.cell(row_idx, 3).value = round(product['sales_2024'], 2)
    product_analysis_sheet.cell(row_idx, 4).value = round(product['growth_rate'], 4)
    product_analysis_sheet.cell(row_idx, 5).value = round(product['actual_margin_rate'], 4)
    product_analysis_sheet.cell(row_idx, 6).value = round(product['margin_amount'], 2)
    product_analysis_sheet.cell(row_idx, 7).value = round(product['sales_ratio'], 4)
    product_analysis_sheet.cell(row_idx, 8).value = product['abc']
    product_analysis_sheet.cell(row_idx, 9).value = product['suggestion']

print(f"产品分析表填充完成，共 {len(product_list)} 个产品")
print("\nABC分类结果：")
for abc_class in ['A类', 'B类', 'C类']:
    products = [p['name'] for p in product_list if p['abc'] == abc_class]
    print(f"  {abc_class}: {', '.join(products)}")

# 保存文件
wb.save('output/result_without_skill_2.xlsx')
print("\n✓ 分析完成！文件已保存到 output/result_without_skill_2.xlsx")
