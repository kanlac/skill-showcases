"""
生成财务报表分析模板文件
包含三张表：损益表、资产负债表、现金流量表
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_financial_template():
    wb = Workbook()

    # 删除默认的 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 创建三张表
    income_stmt = wb.create_sheet("损益表", 0)
    balance_sheet = wb.create_sheet("资产负债表", 1)
    cashflow_stmt = wb.create_sheet("现金流量表", 2)

    # ==================== 损益表 ====================
    income_stmt['A1'] = "损益表 (Income Statement)"
    income_stmt['A1'].font = Font(bold=True, size=14)
    income_stmt.merge_cells('A1:F1')

    # 表头
    income_stmt['A2'] = "项目"
    income_stmt['B2'] = "2022"
    income_stmt['C2'] = "2023"
    income_stmt['D2'] = "2024"
    income_stmt['E2'] = "2025E"
    income_stmt['F2'] = "2026E"

    # 设置表头格式
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = income_stmt[f'{col}2']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    items = [
        ("营业收入", 50000, 60000, 72000, None, None),
        ("营业成本", 30000, 35000, 40000, None, None),
        ("毛利", None, None, None, None, None),  # 公式
        ("销售费用", 5000, 6000, 7200, None, None),
        ("管理费用", 3000, 3500, 4000, None, None),
        ("研发费用", 2000, 2500, 3000, None, None),
        ("营业利润", None, None, None, None, None),  # 公式
        ("财务费用", 500, 600, 700, None, None),
        ("税前利润", None, None, None, None, None),  # 公式
        ("所得税", None, None, None, None, None),  # 公式 (25%税率)
        ("净利润", None, None, None, None, None),  # 公式
    ]

    row = 3
    for item_name, *values in items:
        income_stmt[f'A{row}'] = item_name

        # 填充历史数据 (蓝色表示输入值)
        for col_idx, value in enumerate(values[:3], start=2):
            if value is not None:
                cell = income_stmt.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = Font(color="0000FF")  # 蓝色表示硬编码输入
                cell.number_format = '#,##0'

        row += 1

    # 添加公式 (黑色表示公式)
    # 毛利 = 营业收入 - 营业成本
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt[f'{col}5'] = f'={col}3-{col}4'
        income_stmt[f'{col}5'].number_format = '#,##0'

    # 营业利润 = 毛利 - 销售费用 - 管理费用 - 研发费用
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt[f'{col}9'] = f'={col}5-{col}6-{col}7-{col}8'
        income_stmt[f'{col}9'].number_format = '#,##0'

    # 税前利润 = 营业利润 - 财务费用
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt[f'{col}11'] = f'={col}9-{col}10'
        income_stmt[f'{col}11'].number_format = '#,##0'

    # 所得税 = 税前利润 * 25%
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt[f'{col}12'] = f'={col}11*0.25'
        income_stmt[f'{col}12'].number_format = '#,##0'

    # 净利润 = 税前利润 - 所得税
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt[f'{col}13'] = f'={col}11-{col}12'
        income_stmt[f'{col}13'].number_format = '#,##0'
        income_stmt[f'{col}13'].font = Font(bold=True)

    # 设置列宽
    income_stmt.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        income_stmt.column_dimensions[col].width = 15

    # ==================== 资产负债表 ====================
    balance_sheet['A1'] = "资产负债表 (Balance Sheet)"
    balance_sheet['A1'].font = Font(bold=True, size=14)
    balance_sheet.merge_cells('A1:F1')

    # 表头
    balance_sheet['A2'] = "项目"
    balance_sheet['B2'] = "2022"
    balance_sheet['C2'] = "2023"
    balance_sheet['D2'] = "2024"
    balance_sheet['E2'] = "2025E"
    balance_sheet['F2'] = "2026E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = balance_sheet[f'{col}2']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 资产数据
    assets = [
        ("资产", "", "", "", "", ""),
        ("流动资产", 25000, 30000, 36000, None, None),
        ("  现金及等价物", 10000, 12000, 15000, None, None),
        ("  应收账款", 8000, 10000, 12000, None, None),
        ("  存货", 7000, 8000, 9000, None, None),
        ("非流动资产", 45000, 50000, 54000, None, None),
        ("  固定资产净值", 35000, 38000, 40000, None, None),
        ("  无形资产", 10000, 12000, 14000, None, None),
        ("总资产", None, None, None, None, None),  # 公式
        ("", "", "", "", "", ""),
        ("负债与权益", "", "", "", "", ""),
        ("流动负债", 15000, 18000, 20000, None, None),
        ("  应付账款", 8000, 9000, 10000, None, None),
        ("  短期借款", 7000, 9000, 10000, None, None),
        ("非流动负债", 20000, 22000, 25000, None, None),
        ("  长期借款", 20000, 22000, 25000, None, None),
        ("总负债", None, None, None, None, None),  # 公式
        ("股东权益", None, None, None, None, None),  # 公式
        ("  实收资本", 20000, 20000, 20000, None, None),
        ("  留存收益", None, None, None, None, None),  # 公式
        ("负债与权益总计", None, None, None, None, None),  # 公式
    ]

    row = 3
    for item_name, *values in assets:
        balance_sheet[f'A{row}'] = item_name

        # 填充历史数据
        for col_idx, value in enumerate(values[:3], start=2):
            if value is not None and value != "":
                cell = balance_sheet.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = Font(color="0000FF")
                cell.number_format = '#,##0'

        row += 1

    # 添加公式
    # 总资产 = 流动资产 + 非流动资产
    for col in ['B', 'C', 'D', 'E', 'F']:
        balance_sheet[f'{col}11'] = f'={col}4+{col}9'
        balance_sheet[f'{col}11'].font = Font(bold=True)
        balance_sheet[f'{col}11'].number_format = '#,##0'

    # 总负债 = 流动负债 + 非流动负债
    for col in ['B', 'C', 'D', 'E', 'F']:
        balance_sheet[f'{col}19'] = f'={col}14+{col}17'
        balance_sheet[f'{col}19'].font = Font(bold=True)
        balance_sheet[f'{col}19'].number_format = '#,##0'

    # 留存收益 = 前期留存收益 + 净利润
    balance_sheet['B22'] = 15000  # 2022初始值
    balance_sheet['B22'].font = Font(color="0000FF")
    balance_sheet['B22'].number_format = '#,##0'

    balance_sheet['C22'] = "=B22+损益表!B13"
    balance_sheet['C22'].font = Font(color="00FF00")  # 绿色表示跨表链接
    balance_sheet['C22'].number_format = '#,##0'

    balance_sheet['D22'] = "=C22+损益表!C13"
    balance_sheet['D22'].font = Font(color="00FF00")
    balance_sheet['D22'].number_format = '#,##0'

    balance_sheet['E22'] = "=D22+损益表!D13"
    balance_sheet['E22'].font = Font(color="00FF00")
    balance_sheet['E22'].number_format = '#,##0'

    balance_sheet['F22'] = "=E22+损益表!E13"
    balance_sheet['F22'].font = Font(color="00FF00")
    balance_sheet['F22'].number_format = '#,##0'

    # 股东权益 = 实收资本 + 留存收益
    for col in ['B', 'C', 'D', 'E', 'F']:
        balance_sheet[f'{col}20'] = f'={col}21+{col}22'
        balance_sheet[f'{col}20'].font = Font(bold=True)
        balance_sheet[f'{col}20'].number_format = '#,##0'

    # 负债与权益总计 = 总负债 + 股东权益
    for col in ['B', 'C', 'D', 'E', 'F']:
        balance_sheet[f'{col}23'] = f'={col}19+{col}20'
        balance_sheet[f'{col}23'].font = Font(bold=True)
        balance_sheet[f'{col}23'].number_format = '#,##0'

    # 设置列宽
    balance_sheet.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        balance_sheet.column_dimensions[col].width = 15

    # ==================== 现金流量表 ====================
    cashflow_stmt['A1'] = "现金流量表 (Cash Flow Statement)"
    cashflow_stmt['A1'].font = Font(bold=True, size=14)
    cashflow_stmt.merge_cells('A1:F1')

    # 表头
    cashflow_stmt['A2'] = "项目"
    cashflow_stmt['B2'] = "2022"
    cashflow_stmt['C2'] = "2023"
    cashflow_stmt['D2'] = "2024"
    cashflow_stmt['E2'] = "2025E"
    cashflow_stmt['F2'] = "2026E"

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = cashflow_stmt[f'{col}2']
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 现金流数据
    cashflows = [
        ("经营活动现金流", "", "", "", "", ""),
        ("  净利润", None, None, None, None, None),  # 链接到损益表
        ("  折旧摊销", 3000, 3500, 4000, None, None),
        ("  应收账款变动", None, None, None, None, None),  # 公式
        ("  存货变动", None, None, None, None, None),  # 公式
        ("  应付账款变动", None, None, None, None, None),  # 公式
        ("经营活动现金流净额", None, None, None, None, None),  # 公式
        ("", "", "", "", "", ""),
        ("投资活动现金流", "", "", "", "", ""),
        ("  资本支出", -5000, -6000, -6000, None, None),
        ("投资活动现金流净额", None, None, None, None, None),  # 公式
        ("", "", "", "", "", ""),
        ("筹资活动现金流", "", "", "", "", ""),
        ("  借款增加", 2000, 4000, 4000, None, None),
        ("筹资活动现金流净额", None, None, None, None, None),  # 公式
        ("", "", "", "", "", ""),
        ("现金净增加额", None, None, None, None, None),  # 公式
        ("期初现金", None, None, None, None, None),  # 公式
        ("期末现金", None, None, None, None, None),  # 公式
    ]

    row = 3
    for item_name, *values in cashflows:
        cashflow_stmt[f'A{row}'] = item_name

        # 填充历史数据
        for col_idx, value in enumerate(values[:3], start=2):
            if value is not None and value != "":
                cell = cashflow_stmt.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = Font(color="0000FF")
                cell.number_format = '#,##0'

        row += 1

    # 添加公式
    # 净利润（链接到损益表）
    for col_idx, col in enumerate(['B', 'C', 'D', 'E', 'F'], start=2):
        cashflow_stmt[f'{col}4'] = f'=损益表!{col}13'
        cashflow_stmt[f'{col}4'].font = Font(color="00FF00")
        cashflow_stmt[f'{col}4'].number_format = '#,##0'

    # 应收账款变动（负数表示增加占用现金）
    cashflow_stmt['B6'] = 0  # 第一年无变动
    cashflow_stmt['B6'].font = Font(color="0000FF")
    for col_idx, (col, prev_col) in enumerate(zip(['C', 'D', 'E', 'F'], ['B', 'C', 'D', 'E']), start=3):
        cashflow_stmt[f'{col}6'] = f'=资产负债表!{prev_col}6-资产负债表!{col}6'
        cashflow_stmt[f'{col}6'].font = Font(color="00FF00")
        cashflow_stmt[f'{col}6'].number_format = '#,##0'

    # 存货变动
    cashflow_stmt['B7'] = 0
    cashflow_stmt['B7'].font = Font(color="0000FF")
    for col_idx, (col, prev_col) in enumerate(zip(['C', 'D', 'E', 'F'], ['B', 'C', 'D', 'E']), start=3):
        cashflow_stmt[f'{col}7'] = f'=资产负债表!{prev_col}7-资产负债表!{col}7'
        cashflow_stmt[f'{col}7'].font = Font(color="00FF00")
        cashflow_stmt[f'{col}7'].number_format = '#,##0'

    # 应付账款变动（正数表示增加提供现金）
    cashflow_stmt['B8'] = 0
    cashflow_stmt['B8'].font = Font(color="0000FF")
    for col_idx, (col, prev_col) in enumerate(zip(['C', 'D', 'E', 'F'], ['B', 'C', 'D', 'E']), start=3):
        cashflow_stmt[f'{col}8'] = f'=资产负债表!{col}15-资产负债表!{prev_col}15'
        cashflow_stmt[f'{col}8'].font = Font(color="00FF00")
        cashflow_stmt[f'{col}8'].number_format = '#,##0'

    # 经营活动现金流净额
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt[f'{col}9'] = f'={col}4+{col}5+{col}6+{col}7+{col}8'
        cashflow_stmt[f'{col}9'].font = Font(bold=True)
        cashflow_stmt[f'{col}9'].number_format = '#,##0'

    # 投资活动现金流净额
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt[f'{col}13'] = f'={col}12'
        cashflow_stmt[f'{col}13'].number_format = '#,##0'

    # 筹资活动现金流净额
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt[f'{col}17'] = f'={col}16'
        cashflow_stmt[f'{col}17'].number_format = '#,##0'

    # 现金净增加额
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt[f'{col}19'] = f'={col}9+{col}13+{col}17'
        cashflow_stmt[f'{col}19'].font = Font(bold=True)
        cashflow_stmt[f'{col}19'].number_format = '#,##0'

    # 期初现金
    cashflow_stmt['B20'] = '=资产负债表!B5'
    cashflow_stmt['B20'].font = Font(color="00FF00")
    for col, prev_col in zip(['C', 'D', 'E', 'F'], ['B', 'C', 'D', 'E']):
        cashflow_stmt[f'{col}20'] = f'={prev_col}21'
        cashflow_stmt[f'{col}20'].font = Font(color="00FF00")
        cashflow_stmt[f'{col}20'].number_format = '#,##0'

    # 期末现金（应与资产负债表现金一致）
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt[f'{col}21'] = f'={col}20+{col}19'
        cashflow_stmt[f'{col}21'].font = Font(bold=True)
        cashflow_stmt[f'{col}21'].number_format = '#,##0'

    # 设置列宽
    cashflow_stmt.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        cashflow_stmt.column_dimensions[col].width = 15

    # 保存文件
    wb.save('financial_model_template.xlsx')
    print("模板文件已创建: financial_model_template.xlsx")

if __name__ == "__main__":
    create_financial_template()
