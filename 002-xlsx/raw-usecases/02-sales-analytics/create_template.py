"""
生成销售数据分析模板文件
包含原始销售数据、产品主数据、区域信息等多个数据源
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
from datetime import datetime, timedelta

def create_sales_template():
    wb = Workbook()

    # 删除默认的 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ==================== 1. 原始销售数据 ====================
    sales_data_ws = wb.create_sheet("原始销售数据", 0)

    # 生成 2023-2024 两年的月度销售数据
    products = ['笔记本电脑', '台式电脑', '显示器', '键盘', '鼠标', '耳机', '音箱', '摄像头']
    regions = ['华北', '华东', '华南', '华中', '西南', '西北', '东北']
    channels = ['直销', '经销商', '电商']

    sales_records = []
    start_date = datetime(2023, 1, 1)

    # 为每个产品、区域、渠道生成24个月的数据
    for month in range(24):
        current_date = start_date + timedelta(days=30*month)
        year_month = current_date.strftime('%Y-%m')

        for product in products:
            for region in regions:
                for channel in channels:
                    # 基础销量（带一些随机性和季节性）
                    base_qty = random.randint(50, 200)
                    seasonal_factor = 1.2 if month % 12 in [10, 11] else 1.0  # 11-12月旺季
                    trend_factor = 1 + (month * 0.02)  # 年增长趋势

                    quantity = int(base_qty * seasonal_factor * trend_factor)

                    # 根据产品类型设定基础单价
                    price_map = {
                        '笔记本电脑': 5000,
                        '台式电脑': 4000,
                        '显示器': 1500,
                        '键盘': 200,
                        '鼠标': 100,
                        '耳机': 300,
                        '音箱': 500,
                        '摄像头': 400
                    }
                    unit_price = price_map[product] * (1 + random.uniform(-0.1, 0.1))

                    sales_records.append({
                        '年月': year_month,
                        '产品名称': product,
                        '销售区域': region,
                        '销售渠道': channel,
                        '销售数量': quantity,
                        '单价': round(unit_price, 2),
                        '销售额': round(quantity * unit_price, 2)
                    })

    # 转换为 DataFrame
    df_sales = pd.DataFrame(sales_records)

    # 写入工作表
    sales_data_ws['A1'] = '原始销售数据'
    sales_data_ws['A1'].font = Font(bold=True, size=14)
    sales_data_ws.merge_cells('A1:G1')

    # 表头
    headers = ['年月', '产品名称', '销售区域', '销售渠道', '销售数量', '单价', '销售额']
    for col_idx, header in enumerate(headers, start=1):
        cell = sales_data_ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 写入数据（蓝色表示原始数据）
    for row_idx, row in enumerate(dataframe_to_rows(df_sales, index=False, header=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            cell = sales_data_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(color="0000FF")

            # 数字格式
            if col_idx == 5:  # 销售数量
                cell.number_format = '#,##0'
            elif col_idx in [6, 7]:  # 单价、销售额
                cell.number_format = '#,##0.00'

    # 设置列宽
    sales_data_ws.column_dimensions['A'].width = 12
    sales_data_ws.column_dimensions['B'].width = 15
    sales_data_ws.column_dimensions['C'].width = 12
    sales_data_ws.column_dimensions['D'].width = 12
    sales_data_ws.column_dimensions['E'].width = 12
    sales_data_ws.column_dimensions['F'].width = 12
    sales_data_ws.column_dimensions['G'].width = 15

    # ==================== 2. 产品主数据 ====================
    product_master_ws = wb.create_sheet("产品主数据", 1)

    product_master_data = [
        ['产品名称', '产品类别', '产品代码', '标准成本', '目标毛利率', '是否停产'],
        ['笔记本电脑', '计算机', 'P001', 3500, 0.30, '否'],
        ['台式电脑', '计算机', 'P002', 2800, 0.30, '否'],
        ['显示器', '外设', 'P003', 900, 0.40, '否'],
        ['键盘', '外设', 'P004', 120, 0.40, '否'],
        ['鼠标', '外设', 'P005', 60, 0.40, '否'],
        ['耳机', '音频', 'P006', 180, 0.40, '否'],
        ['音箱', '音频', 'P007', 300, 0.40, '否'],
        ['摄像头', '外设', 'P008', 250, 0.38, '否'],
    ]

    for row_idx, row in enumerate(product_master_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = product_master_ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            if row_idx == 1:  # 表头
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = Font(color="0000FF")

                # 数字格式
                if col_idx == 4:  # 标准成本
                    cell.number_format = '#,##0'
                elif col_idx == 5:  # 目标毛利率
                    cell.number_format = '0.0%'

    # 设置列宽
    product_master_ws.column_dimensions['A'].width = 15
    product_master_ws.column_dimensions['B'].width = 12
    product_master_ws.column_dimensions['C'].width = 12
    product_master_ws.column_dimensions['D'].width = 12
    product_master_ws.column_dimensions['E'].width = 15
    product_master_ws.column_dimensions['F'].width = 12

    # ==================== 3. 区域目标 ====================
    region_target_ws = wb.create_sheet("区域目标", 2)

    region_target_data = [
        ['销售区域', '2023年目标', '2024年目标', '销售经理', '区域等级'],
        ['华北', 15000000, 18000000, '张伟', 'A'],
        ['华东', 20000000, 24000000, '李娜', 'A'],
        ['华南', 18000000, 21600000, '王军', 'A'],
        ['华中', 12000000, 14400000, '刘洋', 'B'],
        ['西南', 10000000, 12000000, '陈晨', 'B'],
        ['西北', 8000000, 9600000, '赵敏', 'C'],
        ['东北', 9000000, 10800000, '孙磊', 'B'],
    ]

    for row_idx, row in enumerate(region_target_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = region_target_ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            if row_idx == 1:  # 表头
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            else:
                if col_idx in [2, 3]:  # 目标金额用黄色背景标注
                    cell.font = Font(color="0000FF")
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.number_format = '#,##0'
                else:
                    cell.font = Font(color="0000FF")

    # 设置列宽
    region_target_ws.column_dimensions['A'].width = 12
    region_target_ws.column_dimensions['B'].width = 15
    region_target_ws.column_dimensions['C'].width = 15
    region_target_ws.column_dimensions['D'].width = 12
    region_target_ws.column_dimensions['E'].width = 12

    # ==================== 4. 月度汇总（需要用公式填充）====================
    monthly_summary_ws = wb.create_sheet("月度汇总", 3)

    monthly_summary_ws['A1'] = '月度销售汇总表（待填充）'
    monthly_summary_ws['A1'].font = Font(bold=True, size=14)
    monthly_summary_ws.merge_cells('A1:H1')

    # 表头
    summary_headers = ['年月', '总销售额', '总销售数量', '平均单价', '同比增长率', '环比增长率', '目标完成率', '备注']
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = monthly_summary_ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 生成年月列表（作为参考）
    months = []
    for year in [2023, 2024]:
        for month in range(1, 13):
            months.append(f'{year}-{month:02d}')

    for row_idx, month in enumerate(months, start=3):
        monthly_summary_ws[f'A{row_idx}'] = month
        monthly_summary_ws[f'A{row_idx}'].font = Font(color="0000FF")

    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        monthly_summary_ws.column_dimensions[col].width = 15

    # ==================== 5. 产品分析（待填充）====================
    product_analysis_ws = wb.create_sheet("产品分析", 4)

    product_analysis_ws['A1'] = '产品销售分析表（待填充）'
    product_analysis_ws['A1'].font = Font(bold=True, size=14)
    product_analysis_ws.merge_cells('A1:I1')

    # 表头
    product_headers = ['产品名称', '2023销售额', '2024销售额', '增长率', '实际毛利率', '毛利额', '销售占比', 'ABC分类', '建议']
    for col_idx, header in enumerate(product_headers, start=1):
        cell = product_analysis_ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # 产品列表
    for row_idx, product in enumerate(products, start=3):
        product_analysis_ws[f'A{row_idx}'] = product
        product_analysis_ws[f'A{row_idx}'].font = Font(color="0000FF")

    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        product_analysis_ws.column_dimensions[col].width = 15

    # 保存文件
    wb.save('sales_analytics_template.xlsx')
    print("销售分析模板文件已创建: sales_analytics_template.xlsx")
    print(f"生成了 {len(sales_records)} 条销售记录")

if __name__ == "__main__":
    # 设置随机种子以保证可重现性
    random.seed(42)
    create_sales_template()
